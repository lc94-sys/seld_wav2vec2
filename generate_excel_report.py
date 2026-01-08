#!/usr/bin/env python3
"""
Generate Excel report from conflict analysis results.
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_report(data: dict, output_path: str):
    """Create formatted Excel report from conflict data."""
    
    wb = Workbook()
    
    # Summary sheet
    summary = wb.active
    summary.title = "Summary"
    
    # Styles
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    high_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    medium_fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
    low_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary sheet content
    summary['A1'] = "Procedure Document Conflict Report"
    summary['A1'].font = title_font
    summary.merge_cells('A1:D1')
    
    summary['A3'] = "Analysis Date:"
    summary['B3'] = data.get('analysis_date', datetime.now().isoformat())
    summary['A4'] = "Documents Analyzed:"
    summary['B4'] = data.get('total_documents', 0)
    summary['A5'] = "Total Conflicts Found:"
    summary['B5'] = data.get('total_conflicts', 0)
    
    severity = data.get('conflicts_by_severity', {})
    summary['A7'] = "Conflicts by Severity"
    summary['A7'].font = Font(bold=True)
    summary['A8'] = "HIGH"
    summary['B8'] = severity.get('HIGH', 0)
    summary['A8'].fill = high_fill
    summary['A8'].font = Font(color="FFFFFF", bold=True)
    
    summary['A9'] = "MEDIUM"
    summary['B9'] = severity.get('MEDIUM', 0)
    summary['A9'].fill = medium_fill
    summary['A9'].font = Font(bold=True)
    
    summary['A10'] = "LOW"
    summary['B10'] = severity.get('LOW', 0)
    summary['A10'].fill = low_fill
    summary['A10'].font = Font(bold=True)
    
    # Set column widths
    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 40
    
    # Conflicts detail sheet
    if data.get('total_conflicts', 0) > 0:
        conflicts = wb.create_sheet("Conflicts Detail")
        
        # Headers
        headers = ["#", "Severity", "Topic", "Document A", "Document A States", 
                   "Document B", "Document B States", "Explanation"]
        for col, header in enumerate(headers, 1):
            cell = conflicts.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Data rows
        for row_idx, conflict in enumerate(data.get('conflicts', []), 2):
            doc_a = conflict.get('document_a', {})
            doc_b = conflict.get('document_b', {})
            severity_val = conflict.get('severity', 'LOW')
            
            row_data = [
                row_idx - 1,
                severity_val,
                conflict.get('topic', ''),
                doc_a.get('name', ''),
                doc_a.get('states', ''),
                doc_b.get('name', ''),
                doc_b.get('states', ''),
                conflict.get('explanation', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = conflicts.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color severity cell
                if col == 2:
                    if severity_val == 'HIGH':
                        cell.fill = high_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif severity_val == 'MEDIUM':
                        cell.fill = medium_fill
                        cell.font = Font(bold=True)
                    elif severity_val == 'LOW':
                        cell.fill = low_fill
                        cell.font = Font(bold=True)
        
        # Column widths
        col_widths = [5, 10, 30, 20, 40, 20, 40, 50]
        for col, width in enumerate(col_widths, 1):
            conflicts.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        conflicts.freeze_panes = 'A2'
        
        # Add autofilter
        conflicts.auto_filter.ref = f"A1:H{len(data.get('conflicts', [])) + 1}"
    
    # HIGH severity sheet (for quick reference)
    high_conflicts = [c for c in data.get('conflicts', []) if c.get('severity') == 'HIGH']
    if high_conflicts:
        high_sheet = wb.create_sheet("HIGH Priority")
        
        headers = ["Topic", "Document A", "What A States", "Document B", "What B States", "Why It Conflicts"]
        for col, header in enumerate(headers, 1):
            cell = high_sheet.cell(row=1, column=col, value=header)
            cell.fill = high_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = border
        
        for row_idx, conflict in enumerate(high_conflicts, 2):
            doc_a = conflict.get('document_a', {})
            doc_b = conflict.get('document_b', {})
            
            row_data = [
                conflict.get('topic', ''),
                doc_a.get('name', ''),
                doc_a.get('states', ''),
                doc_b.get('name', ''),
                doc_b.get('states', ''),
                conflict.get('explanation', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = high_sheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        col_widths = [30, 20, 40, 20, 40, 50]
        for col, width in enumerate(col_widths, 1):
            high_sheet.column_dimensions[get_column_letter(col)].width = width
        
        high_sheet.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"Excel report saved: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_excel_report.py <conflict_report.json> [output.xlsx]")
        sys.exit(1)
    
    json_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "conflict_report.xlsx"
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    create_excel_report(data, output_path)


if __name__ == "__main__":
    main()
